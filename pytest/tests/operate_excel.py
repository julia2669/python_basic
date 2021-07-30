#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @Time : 7/26/2021 11:52 
# @Author : Julia
# @Version：V 0.1
# @File : operate_excel.py
# @desc :

from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

# import warnings
# warnings.filterwarnings('ignore')

# 1. create workbook
wb1 = Workbook()

# 2. load_workbook() 只能处理.xlsx格式的表格
wb=load_workbook('./001.xlsx')

# 3. create sheet
ws1=wb1.create_sheet('ws1')

# 4.current active sheet
ws2=wb1.active
print(ws2.title)

# 5. 指定工作表
ws = wb1['ws1']

# 6. show sheets
print(wb1.sheetnames)

# 7. cell
print(ws.cell(1,1), ws['A1'])

# 8.单元格属性
cell=ws['A1']
print(cell.col_idx)
print(cell.column)
print(cell.row)
print(cell.column_letter)
print(cell.coordinate)
print(cell.data_type)#n :number s:string d:date
print(cell.encoding)
print(cell.has_style, cell.style)
print(cell.style_id)
print(cell.hyperlink)
print(cell.value)
print(ws['A'])
print(ws[1])
print(ws['A:B'])
print(ws[5:10])
print(ws['A3:B9'])
ws['A2']=20
ws.append([1,2,3])#增加一行
print(ws.values) #遍历 返回的是生成器，是将一行数据作为一个元组单元组成的，是由值组成的

for i in ws.values:
    print(i)

for i in ws.iter_rows(min_col=1,max_col=3,min_row=1,max_row=10):
    print(i)

print(ws.rows)#是将一行单元格作为元组单元组成的生成器，与ws.values的区别是，rows返回的是由单元格组成的元组，values是由值组成的

# 注意，删除行或者列后，后面的行或者列会自动往前填充，也就是说，删除第一列，原来的第二列就会变成第一列
ws.delete_cols(1)
ws.delete_rows(1)

# 转pandas
import pandas as pd
df = pd.DataFrame(ws.values)

# pandas 转ws
for i in df.values:
    ws.append(i.tolist())

for i in ws.rows:
    for j in i:
        print(j,j.value,end=',')
    print('')

# 合并单元格
ws.merge_cells("A1:B1")
ws.merge_cells(start_column=3,end_column=6,start_row=2,end_row=3)

# 过滤和排序  实际上，openpyxl可以添加过滤和排序，但是并不会起作用

# 过滤区域
ws.auto_filter.ref = "A:B"
# 给指定列添加过滤条件
ws.auto_filter.add_filter_column(0, ['ASC','DWS'])
ws.auto_filter.add_sort_condition("B2:B15")

# Excel的链接公式

ws['C5'].value = '=HYPERLINK("#Sheet!B2","名称")'

# hyperlink参数

from openpyxl.worksheet.hyperlink import Hyperlink
ws['C6'].hyperlink = Hyperlink(ref='',location='Sheet!H5',target='')
ws['C6'].value = '这是链接'